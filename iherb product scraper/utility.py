import os

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

print("Formatting started!")


def csv_to_xlsx():
    read_file = pd.read_csv('product_details.csv')
    read_file.to_excel('product_details.xlsx', index=None, header=True)


def xlxs_width_formatter():
    workbook = openpyxl.load_workbook("product_details.xlsx")

    worksheet = workbook["Sheet1"]

    MIN_WIDTH = 10
    for i, column_cells in enumerate(worksheet.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        worksheet.column_dimensions[get_column_letter(i)].width = width

    workbook.save(filename='final_product_details.xlsx')


def csv_from_excel():
    read_file = pd.read_excel('final_product_details.xlsx')
    read_file.to_csv('Final.csv', index=False, float_format='%.0f', encoding='utf-8')


if not os.path.exists('final_product_details.xlsx'):
    csv_to_xlsx()
    xlxs_width_formatter()
csv_from_excel()
print("Formatting complete!")
