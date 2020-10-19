"""
A simple script to convert CSV file to xlsx  (without any collapse) 

PERKS:
1. Better formatting
2. File size reduction

"""

import os

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

print("Formatting started!")

INPUT_FILE = input("Enter the name of input file (without extension) (must be CSV) ")
OUTPUT_FILE = input("Enter the destination file name (without extension) (will be xlsx)")


def csv_to_xlsx():
    read_file = pd.read_csv(INPUT_FILE + '.csv')
    read_file.to_excel('temp.xlsx', index=None, header=True)


def xlsx_width_formatter():
    workbook = openpyxl.load_workbook("temp.xlsx")

    worksheet = workbook["Sheet1"]

    MIN_WIDTH = 10
    for i, column_cells in enumerate(worksheet.columns, start=1):
        width = (
            length
            if (length := max(len(str(cell_value) if (cell_value := cell.value) is not None else "")
                              for cell in column_cells)) >= MIN_WIDTH
            else MIN_WIDTH
        )
        worksheet.column_dimensions[get_column_letter(i)].width = width

    workbook.save(filename=OUTPUT_FILE + '.xlsx')


if not os.path.exists(OUTPUT_FILE + '.xlsx'):
    csv_to_xlsx()
    xlsx_width_formatter()

print("Formatting complete!")
